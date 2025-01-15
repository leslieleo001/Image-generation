import os
from datetime import datetime
from pathlib import Path
import pandas as pd
from PyQt6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
    QTableWidget, QTableWidgetItem, QPushButton, 
    QLabel, QFileDialog, QMessageBox, QHeaderView, QMenu,
    QAbstractItemView, QApplication
)
from PyQt6.QtCore import Qt, QSize, QPointF, QPoint
from PyQt6.QtGui import (
    QPixmap, QIcon, QPainter, QPen, QBrush, QColor,
    QCursor, QMouseEvent
)
from src.utils.history_manager import HistoryManager

class DraggableTableWidget(QTableWidget):
    """支持拖放的表格控件"""
    def __init__(self, history_window):
        super().__init__(history_window)
        self.history_manager = None
        self.history_window = history_window
        self.dragging = False
        self.drag_start_pos = None
        self.drag_source_row = -1
        self.drop_indicator_row = -1
        self.drag_pixmap = None
        
        # 设置表格属性
        self.setAcceptDrops(False)  # 禁用Qt的拖放
        self.setDragEnabled(False)
        self.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.setShowGrid(True)
        self.setAlternatingRowColors(True)
        
        # 设置样式
        self.setStyleSheet("""
            QTableWidget {
                gridline-color: #d0d0d0;
            }
            QTableWidget::item:selected {
                background-color: #0078d7;
                color: white;
            }
            QTableWidget::item:hover {
                background-color: #e5f3ff;
            }
        """)
        
        # 启用右键菜单
        self.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.customContextMenuRequested.connect(self.show_context_menu)
        
        # 启用双击事件
        self.cellDoubleClicked.connect(self.handle_double_click)
        
    def mousePressEvent(self, event: QMouseEvent):
        """处理鼠标按下事件"""
        if event.button() == Qt.MouseButton.LeftButton:
            # 获取点击的行
            row = self.rowAt(event.pos().y())
            if row >= 0:
                self.drag_start_pos = event.pos()
                self.drag_source_row = row
                # 选中整行
                self.selectRow(row)
        super().mousePressEvent(event)
        
    def mouseMoveEvent(self, event: QMouseEvent):
        """处理鼠标移动事件"""
        if not (event.buttons() & Qt.MouseButton.LeftButton) or self.drag_source_row < 0:
            return
            
        # 检查是否达到拖动阈值
        if not self.dragging and self.drag_start_pos:
            if (event.pos() - self.drag_start_pos).manhattanLength() < 10:  # 使用固定值替代QApplication.startDragDistance()
                return
                
            self.dragging = True
            # 创建拖动时的预览图像
            row_rect = self.visualRect(self.model().index(self.drag_source_row, 0))
            row_rect.setWidth(self.viewport().width())
            self.drag_pixmap = self.viewport().grab(row_rect)
            
        if self.dragging:
            # 计算目标行
            pos = event.pos()
            row = self.rowAt(pos.y())
            
            # 处理拖到表格边界的情况
            if row < 0:
                if pos.y() < 0:
                    row = 0
                else:
                    row = self.rowCount()
                    
            # 更新拖放指示器位置
            if row != self.drop_indicator_row:
                self.drop_indicator_row = row
                self.viewport().update()
        
    def mouseReleaseEvent(self, event: QMouseEvent):
        """处理鼠标释放事件"""
        if event.button() == Qt.MouseButton.LeftButton and self.dragging:
            try:
                if (self.drop_indicator_row >= 0 and 
                    self.drag_source_row >= 0 and 
                    self.drag_source_row != self.drop_indicator_row and
                    self.drag_source_row < self.rowCount() and
                    self.drop_indicator_row <= self.rowCount()):
                    
                    # 记住当前的滚动位置
                    scrollbar = self.verticalScrollBar()
                    scroll_pos = scrollbar.value()
                    
                    # 移动记录
                    record = self.history_manager.records.pop(self.drag_source_row)
                    insert_pos = self.drop_indicator_row
                    if insert_pos > self.drag_source_row:
                        insert_pos -= 1
                    self.history_manager.records.insert(insert_pos, record)
                    
                    # 保存更改
                    self.history_manager.save_records()
                    
                    # 更新表格内容
                    self.history_window.refresh_table()
                    
                    # 恢复滚动位置并选中移动后的行
                    scrollbar.setValue(scroll_pos)
                    self.selectRow(insert_pos)
                    
                    # 发送历史记录更新信号
                    self.history_manager.history_updated.emit()
                    
            except Exception as e:
                print(f"拖放处理失败: {str(e)}")
                
        # 重置拖动状态
        self.dragging = False
        self.drag_start_pos = None
        self.drag_source_row = -1
        self.drop_indicator_row = -1
        self.drag_pixmap = None
        self.viewport().update()
        
        super().mouseReleaseEvent(event)
        
    def paintEvent(self, event):
        """绘制事件，用于显示拖放指示器和拖动预览"""
        super().paintEvent(event)
        
        if self.dragging and self.drop_indicator_row >= 0:
            painter = QPainter(self.viewport())
            painter.setPen(QPen(QColor("#0078d7"), 2, Qt.PenStyle.SolidLine))
            
            # 计算指示器位置
            if self.drop_indicator_row < self.rowCount():
                rect = self.visualRect(self.model().index(self.drop_indicator_row, 0))
                y = rect.top()
            else:
                rect = self.visualRect(self.model().index(self.rowCount()-1, 0))
                y = rect.bottom()
                
            # 绘制水平线
            width = self.viewport().width()
            painter.drawLine(0, y, width, y)
            
            # 绘制小三角形指示器
            triangle_size = 6
            painter.setBrush(QBrush(QColor("#0078d7")))
            
            # 左侧三角形
            points = [
                QPointF(-triangle_size, y),
                QPointF(0, y - triangle_size),
                QPointF(0, y + triangle_size)
            ]
            painter.drawPolygon(points)
            
            # 右侧三角形
            points = [
                QPointF(width + triangle_size, y),
                QPointF(width, y - triangle_size),
                QPointF(width, y + triangle_size)
            ]
            painter.drawPolygon(points)
            
            # 绘制拖动预览
            if self.drag_pixmap and not self.drag_pixmap.isNull():
                cursor_pos = self.mapFromGlobal(QCursor.pos())
                y = cursor_pos.y() - self.drag_pixmap.height() // 2
                painter.setOpacity(0.7)
                painter.drawPixmap(QPoint(0, y), self.drag_pixmap)

    def handle_double_click(self, row, column):
        """处理双击事件，打开图片"""
        try:
            if row < 0 or row >= len(self.history_manager.records):
                return
                
            record = self.history_manager.records[row]
            if not record:
                return
                
            image_paths = record.get("image_paths", [])
            if not image_paths and "image_path" in record:
                image_paths = [record["image_path"]]
            
            if not image_paths:
                return
                
            # 打开第一张图片
            if os.path.exists(image_paths[0]):
                os.startfile(image_paths[0])
            else:
                QMessageBox.warning(self, "错误", "图片文件不存在")
                
        except Exception as e:
            print(f"打开图片失败: {str(e)}")
            QMessageBox.warning(self, "错误", f"打开图片失败: {str(e)}")

    def show_context_menu(self, pos):
        """显示右键菜单"""
        try:
            menu = QMenu(self)
            
            # 获取当前行
            row = self.indexAt(pos).row()
            if row < 0 or row >= len(self.history_manager.records):
                return
                
            # 获取图片路径
            record = self.history_manager.records[row]
            if not record:
                return
                
            image_paths = record.get("image_paths", [])
            if not image_paths and "image_path" in record:
                image_paths = [record["image_path"]]
            
            if not image_paths:
                return
                
            # 添加菜单项
            open_folder_action = menu.addAction("打开图片所在文件夹")
            
            # 显示菜单并获取选择的动作
            action = menu.exec(self.viewport().mapToGlobal(pos))
            
            if action == open_folder_action:
                # 打开文件夹并选中图片
                if os.path.exists(image_paths[0]):
                    # 获取绝对路径
                    abs_path = os.path.abspath(image_paths[0])
                    # 使用 explorer /select 命令打开文件夹并选中文件
                    os.system(f'explorer /select,"{abs_path}"')
                else:
                    QMessageBox.warning(self, "错误", "文件夹不存在")
                    
        except Exception as e:
            print(f"显示右键菜单失败: {str(e)}")
            QMessageBox.warning(self, "错误", f"显示右键菜单失败: {str(e)}")

class HistoryWindow(QMainWindow):
    def __init__(self, history_manager):
        super().__init__()
        self.history_manager = history_manager
        self.init_ui()
        
    def init_ui(self):
        """初始化界面"""
        self.setWindowTitle("历史记录管理")
        self.resize(1200, 600)
        
        # 创建中心部件
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # 创建布局
        layout = QVBoxLayout()
        central_widget.setLayout(layout)
        
        # 创建工具栏
        toolbar = QHBoxLayout()
        
        # 全选/取消全选按钮
        select_all_btn = QPushButton("全选")
        select_all_btn.clicked.connect(self.select_all_records)
        unselect_all_btn = QPushButton("取消全选")
        unselect_all_btn.clicked.connect(self.unselect_all_records)
        
        # 删除按钮
        delete_btn = QPushButton("删除记录")
        delete_btn.clicked.connect(lambda: self.delete_selected(False))
        delete_with_files_btn = QPushButton("删除记录和文件")
        delete_with_files_btn.clicked.connect(lambda: self.delete_selected(True))
        
        # 导出按钮
        export_btn = QPushButton("导出到Excel")
        export_btn.clicked.connect(self.export_to_excel)
        
        # 刷新按钮
        refresh_btn = QPushButton("刷新")
        refresh_btn.clicked.connect(self.refresh_table)
        
        # 添加按钮到工具栏
        toolbar.addWidget(select_all_btn)
        toolbar.addWidget(unselect_all_btn)
        toolbar.addWidget(delete_btn)
        toolbar.addWidget(delete_with_files_btn)
        toolbar.addWidget(export_btn)
        toolbar.addWidget(refresh_btn)
        toolbar.addStretch()
        
        # 创建表格
        self.table = DraggableTableWidget(self)
        self.table.history_manager = self.history_manager
        
        # 设置表格属性
        self.table.setColumnCount(7)  # 减少一列，移除操作列
        self.table.setHorizontalHeaderLabels([
            "选择", "缩略图", "名称", "提示词", "模型", "参数", "保存路径"
        ])
        
        # 设置表格选择模式
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QTableWidget.SelectionMode.ExtendedSelection)
        
        # 设置列宽
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)  # 复选框列
        self.table.setColumnWidth(0, 30)  # 复选框列宽
        
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Fixed)  # 预览列
        self.table.setColumnWidth(1, 54)  # 预览列宽 (50像素缩略图 + 2个边距 * 2像素)
        
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Interactive)  # 名称列
        self.table.setColumnWidth(2, 150)  # 名称列默认宽度
        
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)  # 提示词列
        
        self.table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.Interactive)  # 模型列
        self.table.setColumnWidth(4, 100)  # 模型列默认宽度
        
        self.table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeMode.Interactive)  # 参数列
        self.table.setColumnWidth(5, 120)  # 参数列默认宽度
        
        self.table.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeMode.Interactive)  # 保存路径列
        self.table.setColumnWidth(6, 200)  # 保存路径列默认宽度
        
        # 添加到布局
        layout.addLayout(toolbar)
        layout.addWidget(self.table)
        
        # 刷新表格
        self.refresh_table()
        
    def select_all_records(self):
        """全选所有记录"""
        for row in range(self.table.rowCount()):
            item = self.table.item(row, 0)
            if item:
                item.setCheckState(Qt.CheckState.Checked)
                
    def unselect_all_records(self):
        """取消全选"""
        for row in range(self.table.rowCount()):
            item = self.table.item(row, 0)
            if item:
                item.setCheckState(Qt.CheckState.Unchecked)
    
    def get_checked_rows(self):
        """获取所有选中的行"""
        checked_rows = []
        for row in range(self.table.rowCount()):
            item = self.table.item(row, 0)
            if item and item.checkState() == Qt.CheckState.Checked:
                checked_rows.append(row)
        return checked_rows
    
    def refresh_table(self):
        """刷新表格数据"""
        records = self.history_manager.get_records()
        
        # 先设置正确的行数
        self.table.setRowCount(len(records))
        
        # 设置统一的行高
        for row in range(len(records)):
            self.table.setRowHeight(row, 54)  # 设置行高为54像素（50像素缩略图 + 2*2像素边距）
        
        # 更新表格内容
        for idx, record in enumerate(records):
            # 添加复选框
            checkbox_item = QTableWidgetItem()
            checkbox_item.setCheckState(Qt.CheckState.Unchecked)
            self.table.setItem(idx, 0, checkbox_item)
            
            # 缩略图
            image_paths = record.get("image_paths", [])  # 获取所有图片路径
            if not image_paths and "image_path" in record:  # 兼容旧格式
                image_paths = [record["image_path"]]
            
            if image_paths:
                # 创建缩略图容器
                thumb_widget = QWidget()
                thumb_layout = QHBoxLayout(thumb_widget)
                thumb_layout.setContentsMargins(2, 2, 2, 2)
                thumb_layout.setSpacing(0)
                
                # 只显示第一张图片的缩略图
                path = image_paths[0]
                if os.path.exists(path):
                    pixmap = QPixmap(path)
                    scaled_pixmap = pixmap.scaled(50, 50, Qt.AspectRatioMode.KeepAspectRatio)
                    label = QLabel()
                    label.setFixedSize(50, 50)
                    label.setAlignment(Qt.AlignmentFlag.AlignCenter)
                    label.setPixmap(scaled_pixmap)
                    thumb_layout.addWidget(label)
                
                self.table.setCellWidget(idx, 1, thumb_widget)
            
            # 名称（使用文件名，不带扩展名）
            if image_paths:
                filename = os.path.splitext(os.path.basename(image_paths[0]))[0]
                if len(image_paths) > 1:
                    filename += f" (+{len(image_paths)-1})"
                self.table.setItem(idx, 2, QTableWidgetItem(filename))
            else:
                self.table.setItem(idx, 2, QTableWidgetItem("未知"))
            
            # 提示词
            params = record.get("params", {})
            prompt = params.get("prompt", "")
            self.table.setItem(idx, 3, QTableWidgetItem(prompt))
            
            # 模型
            model = params.get("model", "")
            self.table.setItem(idx, 4, QTableWidgetItem(model))
            
            # 参数
            param_text = f"尺寸: {params.get('size', '')}\n"
            param_text += f"步数: {params.get('num_inference_steps', '')}\n"
            param_text += f"引导系数: {params.get('guidance_scale', '')}\n"
            param_text += f"数量: {len(image_paths)}"
            self.table.setItem(idx, 5, QTableWidgetItem(param_text))
            
            # 保存路径
            path_text = "\n".join(image_paths)
            self.table.setItem(idx, 6, QTableWidgetItem(path_text))
    
    def delete_selected(self, delete_files=False):
        """删除选中的记录"""
        # 获取选中的行
        selected_rows = set()
        
        # 获取通过点击选中的行
        for item in self.table.selectedItems():
            selected_rows.add(item.row())
        
        # 获取通过复选框选中的行
        for row in range(self.table.rowCount()):
            if self.table.item(row, 0) and self.table.item(row, 0).checkState() == Qt.CheckState.Checked:
                selected_rows.add(row)
        
        selected_rows = sorted(selected_rows, reverse=True)  # 从后往前排序
        
        if not selected_rows:
            QMessageBox.warning(self, "警告", "请先选择要删除的记录")
            return
            
        # 确认删除
        msg = "确定要删除选中的记录吗？"
        if delete_files:
            msg = "确定要删除选中的记录及其关联的图片文件吗？"
        
        reply = QMessageBox.question(
            self, 
            "确认删除",
            msg,
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            try:
                # 从后往前删除，避免索引变化
                for row in selected_rows:
                    if row < len(self.history_manager.records):
                        record = self.history_manager.records[row]
                        
                        # 如果需要删除文件
                        if delete_files:
                            image_paths = record.get("image_paths", [])
                            if not image_paths and "image_path" in record:
                                image_paths = [record["image_path"]]
                            
                            for path in image_paths:
                                if os.path.exists(path):
                                    try:
                                        os.remove(path)
                                    except Exception as e:
                                        print(f"删除文件失败: {path}, 错误: {str(e)}")
                        
                        # 删除记录
                        del self.history_manager.records[row]
                
                # 保存更改
                self.history_manager.save_records()
                
                # 刷新表格
                self.refresh_table()
                
                # 发送历史记录更新信号
                self.history_manager.history_updated.emit()
                
            except Exception as e:
                print(f"删除记录失败: {str(e)}")
                raise  # 重新抛出异常以便测试捕获
    
    def export_to_excel(self):
        """导出选中记录为Excel（包含原图）"""
        try:
            # 获取选中的行
            selected_rows = self.get_checked_rows()
            if not selected_rows:
                QMessageBox.warning(self, "提示", "请先选择要导出的记录")
                return
            
            # 选择保存路径
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "导出Excel",
                "",
                "Excel Files (*.xlsx)"
            )
            
            if not file_path:
                return
                
            # 准备数据
            records = self.history_manager.get_records()
            selected_records = [records[row] for row in selected_rows]
            
            # 创建一个新的Excel工作簿
            from openpyxl import Workbook
            from openpyxl.drawing.image import Image
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            ws = wb.active
            
            # 设置列标题
            headers = [
                "图片", "名称", "提示词", "负面提示词", "模型", 
                "尺寸", "步数", "引导系数", "随机种子", "提示增强", 
                "图片路径"
            ]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # 设置第一列的宽度（用于显示图片）
            ws.column_dimensions['A'].width = 20  # 设置为20个字符宽度（约120像素）
            
            # 设置行高
            ws.row_dimensions[1].height = 15  # 标题行高度
            
            # 填充数据
            for row_idx, record in enumerate(selected_records, 2):  # 从第2行开始（第1行是标题）
                params = record.get("params", {})
                image_paths = record.get("image_paths", [])
                if not image_paths and "image_path" in record:  # 兼容旧格式
                    image_paths = [record["image_path"]]
                
                # 处理图片
                if image_paths:
                    try:
                        # 获取第一张图片
                        img_path = image_paths[0]
                        if os.path.exists(img_path):
                            # 直接使用原图
                            img = Image(img_path)
                            
                            # 计算合适的显示大小
                            # Excel中1个字符宽度约等于6像素，1个单位高度约等于1.5像素
                            max_width = 120  # 20个字符宽度 * 6像素
                            max_height = 100  # 约67个单位高度（100像素）
                            
                            # 计算缩放比例
                            width_scale = max_width / img.width if img.width > max_width else 1
                            height_scale = max_height / img.height if img.height > max_height else 1
                            scale = min(width_scale, height_scale)
                            
                            # 应用缩放
                            img.width = int(img.width * scale)
                            img.height = int(img.height * scale)
                            
                            # 设置行高，确保图片完整显示
                            # 将像素转换为Excel单位（1像素约等于0.75单位）
                            row_height = img.height * 0.75
                            # 确保最小行高
                            row_height = max(row_height, 20)
                            ws.row_dimensions[row_idx].height = row_height
                            
                            # 将图片添加到单元格，并设置锚点使其居中
                            ws.add_image(img, f'A{row_idx}')
                    except Exception as e:
                        print(f"处理图片失败: {str(e)}")
                
                # 获取文件名（不带扩展名）
                filename = "未知"
                if image_paths:
                    filename = os.path.splitext(os.path.basename(image_paths[0]))[0]
                    if len(image_paths) > 1:
                        filename += f" (+{len(image_paths)-1})"
                
                # 填充其他列
                row_data = [
                    "",  # 第一列是图片
                    filename,  # 名称列
                    params.get("prompt", ""),
                    params.get("negative_prompt", ""),
                    params.get("model", ""),
                    params.get("size", ""),
                    params.get("num_inference_steps", ""),
                    params.get("guidance_scale", ""),
                    params.get("seed", ""),
                    "是" if params.get("prompt_enhancement", False) else "否",
                    "\n".join(image_paths)
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    # 对于长文本，启用自动换行
                    if isinstance(value, str) and len(value) > 50:
                        cell.alignment = cell.alignment.copy(wrapText=True)
            
            # 调整列宽
            for col in range(2, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 20
            
            # 保存Excel文件
            wb.save(file_path)
            
            QMessageBox.information(self, "提示", f"已导出 {len(selected_records)} 条记录")
            
        except Exception as e:
            QMessageBox.warning(self, "错误", f"导出失败: {str(e)}") 